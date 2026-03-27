"""
Upload Excel data to Supabase
Run: pip install supabase openpyxl && python upload_to_supabase.py
"""

from supabase import create_client
import openpyxl
from datetime import datetime

SUPABASE_URL = "https://zovnmmdfthpbubrorsgh.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inpvdm5tbWRmdGhwYnVicm9yc2doIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjE1NzE3ODgsImV4cCI6MjA3NzE0Nzc4OH0.92BH2sjUOgkw6iSRj1_4gt0p3eThg3QT4VK-Q4EdmBE"
EXCEL_PATH = "/Users/raghunandanmali/Downloads/RAMRAJ Report 2026-27.xlsx"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Load workbook with computed values
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ==================== UPLOAD DAILY SALES ====================
print("Uploading Daily Sales...")
ws = wb["Daily sales"]

daily_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    date_val = row[0]
    sales_val = row[1]

    # Skip rows with no sales data
    if not date_val or sales_val is None:
        continue

    date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)

    entry = {
        "date": date_str,
        "sales": float(sales_val or 0),
        "no_of_bills": int(row[2] or 0),
        "sales_returns": float(row[3] or 0),
        "net_sales": float(row[4] or 0),
        "closing_stock": float(row[5] or 0),
        "bank_balance": float(row[6] or 0),
        "orders_placed_amount": float(row[7] or 0),
        "salary_paid": float(row[8] or 0),
        "electricity_paid": float(row[9] or 0),
        "admin_expenses": float(row[10] or 0),
        "total_expenses": float(row[11] or 0),
    }
    daily_rows.append(entry)

if daily_rows:
    result = supabase.table("daily_sales").upsert(daily_rows, on_conflict="date").execute()
    print(f"  Uploaded {len(daily_rows)} daily sales records")
else:
    print("  No daily sales data found")

# ==================== UPLOAD CONSOLIDATED ====================
print("Uploading Monthly Consolidated...")
ws2 = wb["Consolidated "]

monthly_rows = []
for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
    month_val = row[0]
    sales_val = row[2]

    if not month_val or sales_val is None:
        continue

    # Handle string months like 'Sept-25'
    if isinstance(month_val, str):
        if month_val.startswith("Sept"):
            month_str = "2025-09-01"
        elif month_val.startswith("Closing"):
            continue
        else:
            continue
    elif isinstance(month_val, datetime):
        month_str = month_val.strftime("%Y-%m-01")
    else:
        continue

    entry = {
        "month": month_str,
        "no_of_days": int(row[1] or 0),
        "sales": float(row[2] or 0),
        "sales_return": float(row[3] or 0),
        "net_sales": float(row[4] or 0),
        "avg_sales_monthly": float(row[5] or 0),
    }
    monthly_rows.append(entry)

if monthly_rows:
    result = supabase.table("monthly_consolidated").upsert(monthly_rows, on_conflict="month").execute()
    print(f"  Uploaded {len(monthly_rows)} monthly records")
else:
    print("  No consolidated data found")

print("\nDone! Data uploaded to Supabase successfully.")
